from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
import copy
from pathlib import Path

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────
NAVY      = "0D1B3E"
GOLD      = "C9A84C"
WHITE     = "FFFFFF"
LGOLD     = "F5F0E0"
LNAVY     = "E8ECF5"
GREEN     = "1E7E34"
LGREEN    = "D4EDDA"
ORANGE    = "D4720A"
LORANGE   = "FFF3CD"
RED       = "C0392B"
LRED      = "FADBD8"
GRAY      = "6C757D"
LGRAY     = "F8F9FA"
MID       = "DEE2E6"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False, name="Arial"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(all_sides="thin", color="CCCCCC"):
    s = Side(style=all_sides, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom(color=NAVY):
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def apply(ws, cell_ref, value=None, bg=None, fg=None, bold=False, size=10,
          italic=False, h_align="left", v_align="center", wrap=False,
          bdr=None, number_format=None):
    cell = ws[cell_ref]
    if value is not None:
        cell.value = value
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, size=size, color=fg or "000000", italic=italic)
    cell.alignment = align(h_align, v_align, wrap)
    if bdr:
        cell.border = bdr
    if number_format:
        cell.number_format = number_format
    return cell

def section_header(ws, row, col, text, bg=NAVY, fg=WHITE, size=12, merge_to=None, h="left"):
    ref = f"{get_column_letter(col)}{row}"
    ws[ref].value = text
    ws[ref].fill = fill(bg)
    ws[ref].font = font(bold=True, size=size, color=fg)
    ws[ref].alignment = align(h, "center")
    if merge_to:
        end = f"{get_column_letter(merge_to)}{row}"
        ws.merge_cells(f"{ref}:{end}")
    return row + 1

def row_label(ws, row, col, text, bg=LGRAY, bold=False, size=10):
    ref = f"{get_column_letter(col)}{row}"
    ws[ref].value = text
    ws[ref].fill = fill(bg)
    ws[ref].font = font(bold=bold, size=size, color=NAVY)
    ws[ref].alignment = align("left", "center", wrap=True)
    ws[ref].border = border()

def data_cell(ws, row, col, value, bg=WHITE, fg="333333", bold=False,
              wrap=True, size=10, bdr=True):
    ref = f"{get_column_letter(col)}{row}"
    ws[ref].value = value
    ws[ref].fill = fill(bg)
    ws[ref].font = font(bold=bold, size=size, color=fg)
    ws[ref].alignment = align("left", "center", wrap=wrap)
    if bdr:
        ws[ref].border = border()
    return ws[ref]

def set_col_width(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def set_row_height(ws, heights_dict):
    for r, h in heights_dict.items():
        ws.row_dimensions[r].height = h

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ══════════════════════════════════════════════════════════════
#  SHEET 1: FUNNEL MAP (Overview + Workflow Diagram)
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Funnel Map"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = NAVY

set_col_width(ws1, {
    "A":2,"B":18,"C":2,"D":20,"E":2,"F":22,"G":2,"H":22,"I":2,"J":18,"K":2
})

# Title block
for col in range(1, 12):
    ws1.cell(row=1, column=col).fill = fill(NAVY)
    ws1.cell(row=2, column=col).fill = fill(NAVY)
    ws1.cell(row=3, column=col).fill = fill(NAVY)

ws1.merge_cells("B1:J3")
ws1["B1"].value = "EVERMORE LIFE  —  META FINAL EXPENSE FUNNEL MAP"
ws1["B1"].font = Font(bold=True, size=20, color=GOLD, name="Arial")
ws1["B1"].alignment = align("center", "center")
ws1.row_dimensions[1].height = 15
ws1.row_dimensions[2].height = 30
ws1.row_dimensions[3].height = 15

# Funnel stage boxes (text-based diagram)
stages = [
    # (start_row, label_row_range, box_col, color, title, bullets)
    (5,  8,  "B", NAVY,   GOLD,   "STAGE 1\nFACEBOOK ADS",
     ["• Final Expense video ad (60–90 sec)\n• Image/carousel ads\n• Lead Gen form ads\n• Target: Ages 50–75 | Income $25k–$75k\n• Interests: Seniors, AARP, Retirement, Burial"]),
    (10, 13, "B", "1A4A8A", WHITE, "STAGE 2\nTRAFFIC SPLIT",
     ["Path A → Facebook Lead Ad → GHL directly\nPath B → Click-through → Landing page (Sarah)\nPath C → Video views → Retargeting audiences"]),
    (15, 18, "B", "155724", WHITE, "STAGE 3\nSARAH AI ASSISTANT",
     ["• Visitor enters virtual office\n• Sarah chats: qualifies lead\n• Collects: age, health, coverage goal\n• Books appointment → GHL CRM"]),
    (20, 23, "B", ORANGE,  WHITE, "STAGE 4\nGHL CRM + AUTOMATION",
     ["• Contact created automatically\n• SMS confirmation sent to lead\n• Keenan notified instantly\n• Pipeline stage: New Lead"]),
    (25, 28, "B", RED,     WHITE, "STAGE 5\nHUMAN TAKEOVER",
     ["• Licensed agent reviews lead in GHL\n• Calls lead at appointed time\n• Runs carriers + health questions\n• Closes policy"]),
]

def draw_stage_box(ws, start_row, end_row, start_col_idx, bg, title_color, title, bullet_text):
    end_col = start_col_idx + 1
    sc = get_column_letter(start_col_idx)
    ec = get_column_letter(end_col)
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=start_col_idx).fill = fill(bg)
        ws.cell(row=r, column=end_col).fill = fill(bg)
    ws.merge_cells(f"{sc}{start_row}:{ec}{start_row}")
    ws[f"{sc}{start_row}"].value = title
    ws[f"{sc}{start_row}"].font = Font(bold=True, size=11, color=title_color, name="Arial")
    ws[f"{sc}{start_row}"].alignment = align("center", "center", wrap=True)
    ws.row_dimensions[start_row].height = 32
    if end_row > start_row + 1:
        ws.merge_cells(f"{sc}{start_row+1}:{ec}{end_row}")
        ws[f"{sc}{start_row+1}"].value = bullet_text[0]
        ws[f"{sc}{start_row+1}"].font = Font(size=9, color=WHITE, name="Arial")
        ws[f"{sc}{start_row+1}"].alignment = align("left", "top", wrap=True)
        for r in range(start_row + 1, end_row + 1):
            ws.row_dimensions[r].height = 18

for s in stages:
    draw_stage_box(ws1, s[0], s[1], 2, s[3], s[4], s[5], s[6])

# Arrows between stages
arrow_rows = [9, 14, 19, 24]
for r in arrow_rows:
    ws1.row_dimensions[r].height = 14
    ws1.merge_cells(f"B{r}:C{r}")
    ws1[f"B{r}"].value = "▼"
    ws1[f"B{r}"].font = Font(size=14, color=GOLD, bold=True)
    ws1[f"B{r}"].alignment = align("center", "center")

# Retargeting audiences panel (right side)
ret_start = 5
ws1.merge_cells(f"F{ret_start}:J{ret_start}")
ws1[f"F{ret_start}"].value = "RETARGETING AUDIENCE LAYERS"
ws1[f"F{ret_start}"].fill = fill(NAVY)
ws1[f"F{ret_start}"].font = Font(bold=True, size=12, color=GOLD, name="Arial")
ws1[f"F{ret_start}"].alignment = align("center", "center")
ws1.row_dimensions[ret_start].height = 28

ret_layers = [
    ("VIDEO: 15-sec viewers",  "15S AUDIENCE",  LGOLD,  NAVY,
     "They showed interest but left early.\nAd served: Pick up at 15-sec mark. Reinforce the PROBLEM — burial costs, leaving family with debt. New hook."),
    ("VIDEO: 50% viewers",     "50% AUDIENCE",  LORANGE, "7B3F00",
     "They're interested. Mid-funnel.\nAd served: Pick up at 50% — explain the SOLUTION. 'Here's how final expense works. No exam. Locked-in rate.'"),
    ("VIDEO: 75-90% viewers",  "75% AUDIENCE",  LGREEN,  "155724",
     "High intent. Almost convinced.\nAd served: Social proof + urgency. Testimonial or agent quote. 'Don't wait — rates go up with age.'"),
    ("VIDEO: 100% viewers +\nWebsite visitors", "HOT AUDIENCE", LRED, RED,
     "They know everything. Ready.\nAd served: Direct ask — 'Talk to a real person.' Book a call. Introduce real agent. Limited spots framing."),
    ("Lead ad submitters /\nSarah chat visitors", "WARMEST",  LNAVY, NAVY,
     "Already in funnel.\nAd served: Reminder/follow-up. 'We saved your spot.' Or testimonial. Push to complete booking if abandoned."),
]

for i, (label, badge, bg, fg_color, desc) in enumerate(ret_layers):
    r = ret_start + 1 + (i * 4)
    ws1.row_dimensions[r].height = 20
    ws1.row_dimensions[r+1].height = 18
    ws1.row_dimensions[r+2].height = 18
    ws1.row_dimensions[r+3].height = 8

    ws1.merge_cells(f"F{r}:G{r}")
    ws1[f"F{r}"].value = label
    ws1[f"F{r}"].fill = fill(NAVY)
    ws1[f"F{r}"].font = Font(bold=True, size=9, color=WHITE, name="Arial")
    ws1[f"F{r}"].alignment = align("left", "center")

    ws1[f"H{r}"].value = badge
    ws1[f"H{r}"].fill = fill(bg)
    ws1[f"H{r}"].font = Font(bold=True, size=8, color=fg_color, name="Arial")
    ws1[f"H{r}"].alignment = align("center", "center")

    ws1.merge_cells(f"F{r+1}:J{r+2}")
    ws1[f"F{r+1}"].value = desc
    ws1[f"F{r+1}"].fill = fill(bg)
    ws1[f"F{r+1}"].font = Font(size=9, color=fg_color, name="Arial")
    ws1[f"F{r+1}"].alignment = align("left", "top", wrap=True)
    ws1[f"F{r+1}"].border = border()

# GHL + AI label connecting
connect_row = 30
ws1.merge_cells(f"D{connect_row}:J{connect_row}")
ws1[f"D{connect_row}"].value = "⬅  ALL PATHS FEED INTO  →   GoHighLevel CRM  +  Sarah AI  +  Licensed Agent Close"
ws1[f"D{connect_row}"].fill = fill(GOLD)
ws1[f"D{connect_row}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
ws1[f"D{connect_row}"].alignment = align("center", "center")
ws1.row_dimensions[connect_row].height = 22

freeze(ws1, "A4")


# ══════════════════════════════════════════════════════════════
#  SHEET 2: MARKETING STRATEGY
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Marketing Strategy")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = GOLD

set_col_width(ws2, {"A":3,"B":24,"C":60,"D":3})
set_row_height(ws2, {1:12, 2:36, 3:12})

ws2.merge_cells("B2:C2")
ws2["B2"].value = "EVERMORE LIFE — META FINAL EXPENSE MARKETING STRATEGY"
ws2["B2"].fill = fill(NAVY)
ws2["B2"].font = Font(bold=True, size=16, color=GOLD, name="Arial")
ws2["B2"].alignment = align("center", "center")

strategy_sections = [
    ("OBJECTIVE", [
        ("Campaign Goal", "Generate final expense insurance leads aged 50–75 via Meta (Facebook/Instagram). Qualify via Sarah AI. Close with licensed human agents."),
        ("Success Metric", "Cost Per Lead (CPL) under $25. Appointment show rate above 40%. Policy close rate above 20%."),
        ("Budget Approach", "Start: $30–50/day. Scale winning ad sets 2x every 5 days once CPL stabilizes. Pause any ad set with 0 leads after $50 spend."),
    ]),
    ("TARGET AUDIENCE", [
        ("Primary Audience", "Ages 50–75 | United States | Income $25k–$80k | Homeowners or renters\nInterests: AARP, Seniors, Retirement planning, Funeral planning, Medicare, Social Security"),
        ("Secondary Audience", "Children of seniors (ages 35–55) researching coverage for a parent\nInterests: Elder care, Assisted living, Family financial planning"),
        ("Exclusions", "Exclude: Under 45 | Existing customers (email list custom audience) | People who clicked 'Not Interested'"),
        ("Custom Audiences", "Build from: Page engagers | Video viewers (15s, 50%, 75%, 100%) | Lead form submitters | Website visitors (pixel) | Sarah chat completers"),
        ("Lookalike Audiences", "Once 100+ leads collected: build 1% LAL from lead form submitters and policy closers. Scale into 2% and 3% as data grows."),
    ]),
    ("AD PLATFORM SETTINGS", [
        ("Campaign Type", "Awareness (video views) + Leads (lead gen forms) + Traffic (to Sarah landing page)\nRun all three. Video builds audiences. Lead forms convert fast. Traffic feeds Sarah."),
        ("Ad Placements", "Facebook Feed (primary) | Facebook Reels | Instagram Feed | Instagram Reels\nTurn OFF: Audience Network, Messenger (too low intent for insurance)"),
        ("Bidding Strategy", "Start: Lowest cost (let Meta learn). After 50 conversions: switch to Cost Cap at your target CPL."),
        ("Campaign Structure", "1 Campaign = 1 objective\n3 Ad Sets per campaign (broad, interest-stacked, retargeting)\n2-3 creative variations per ad set"),
        ("Pixel & Tracking", "Install Meta Pixel on Sarah's page. Track: PageView, Lead (form submit), Schedule (booking confirm). Use Conversions API via GHL for server-side tracking."),
    ]),
    ("MESSAGING FRAMEWORK", [
        ("Core Hook", "\"Did you know your family could be left with $15,000+ in burial costs if something happened to you tomorrow?\""),
        ("Unique Mechanism", "We're licensed in your state. We shop ALL the top carriers. We find the lowest rate your health qualifies for. No medical exam required for most plans."),
        ("Proof / Trust", "Licensed agents | Multiple carriers | No exam options | Rate locks in for life | Coverage from day 1 for accidental death"),
        ("Objection Pre-Handle", "\"Too expensive?\" — Plans start under $1/day for many ages.\n\"Already have coverage?\" — We often find better rates.\n\"Don't qualify?\" — Guaranteed issue available for most."),
        ("Call to Action",
         "Top of funnel: \"Watch to see if you qualify\"\nMid funnel: \"See how much coverage you can get\"\nBottom funnel: \"Book your free 15-min consultation\""),
    ]),
    ("COMPLIANCE NOTES", [
        ("Required Disclosures", "Always include: 'Not affiliated with any government program.' 'Coverage amounts vary by age and health.' 'Licensed insurance agent.'"),
        ("Prohibited Claims", "NEVER promise specific coverage amounts. NEVER say 'guaranteed approval' without the actual product being guaranteed issue. NEVER use fear in a misleading way."),
        ("Facebook Ad Policy", "Insurance ads require pre-approval in some regions. Use factual language. Avoid before/after framing. No exaggerated claims about savings."),
    ]),
]

current_row = 5
for section_title, rows in strategy_sections:
    ws2.merge_cells(f"B{current_row}:C{current_row}")
    ws2[f"B{current_row}"].value = section_title
    ws2[f"B{current_row}"].fill = fill(NAVY)
    ws2[f"B{current_row}"].font = Font(bold=True, size=11, color=GOLD, name="Arial")
    ws2[f"B{current_row}"].alignment = align("left", "center")
    ws2.row_dimensions[current_row].height = 22
    current_row += 1

    for label, content in rows:
        ws2[f"B{current_row}"].value = label
        ws2[f"B{current_row}"].fill = fill(LNAVY)
        ws2[f"B{current_row}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
        ws2[f"B{current_row}"].alignment = align("left", "top", wrap=True)
        ws2[f"B{current_row}"].border = border()

        ws2[f"C{current_row}"].value = content
        ws2[f"C{current_row}"].fill = fill(WHITE)
        ws2[f"C{current_row}"].font = Font(size=10, color="333333", name="Arial")
        ws2[f"C{current_row}"].alignment = align("left", "top", wrap=True)
        ws2[f"C{current_row}"].border = border()

        line_count = content.count("\n") + 1
        ws2.row_dimensions[current_row].height = max(30, line_count * 16)
        current_row += 1

    current_row += 1

freeze(ws2, "B3")


# ══════════════════════════════════════════════════════════════
#  SHEET 3: AD SETS
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Ad Sets")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = "1A4A8A"

set_col_width(ws3, {
    "A":3,"B":22,"C":18,"D":18,"E":18,"F":18,"G":3
})

ws3.merge_cells("B2:F2")
ws3["B2"].value = "EVERMORE LIFE — META AD SET STRUCTURE"
ws3["B2"].fill = fill(NAVY)
ws3["B2"].font = Font(bold=True, size=14, color=GOLD, name="Arial")
ws3["B2"].alignment = align("center", "center")
ws3.row_dimensions[2].height = 32

headers = ["AD SET NAME", "AUDIENCE", "OBJECTIVE / BID", "BUDGET / SCHEDULE", "CREATIVE ASSIGNED"]
header_row = 4
for i, h in enumerate(headers):
    col = get_column_letter(i + 2)
    ws3[f"{col}{header_row}"].value = h
    ws3[f"{col}{header_row}"].fill = fill(NAVY)
    ws3[f"{col}{header_row}"].font = Font(bold=True, size=10, color=WHITE, name="Arial")
    ws3[f"{col}{header_row}"].alignment = align("center", "center")
    ws3[f"{col}{header_row}"].border = border()
ws3.row_dimensions[header_row].height = 24

ad_sets = [
    # TOP OF FUNNEL
    ("── TOP OF FUNNEL ──", "", "", "", "", NAVY, GOLD),
    ("TOF-1\nBroad 50–75\n(Cold)", "Age 50–75 | USA\nNo interest targeting\nBroad = let Meta find buyers",
     "Video Views\n(ThruPlay)\nLowest Cost",
     "$15/day\nAlways On\nStart here",
     "Video Ad A (Hero)\nThe full 60-sec story", LGOLD, NAVY),
    ("TOF-2\nInterest Stack\n(Cold)", "Age 50–75 | USA\nInterests: AARP, Medicare,\nRetirement, Burial, Seniors",
     "Lead Generation\n(Lead Form)\nLowest Cost",
     "$15/day\nAlways On",
     "Video Ad A + Lead Form\nImage Ad A", LGOLD, NAVY),
    ("TOF-3\nChildren of Seniors\n(Cold)", "Age 35–55 | USA\nInterests: Elder care,\nAssisted living, Family finance",
     "Lead Generation\n(Lead Form)\nLowest Cost",
     "$10/day\nTest 7 days",
     "Image Ad B\n(Angle: 'For your parent')", LGOLD, NAVY),

    # MIDDLE OF FUNNEL — RETARGETING
    ("── MIDDLE OF FUNNEL — RETARGETING ──", "", "", "", "", "1A4A8A", WHITE),
    ("MOF-1\n15-Second Viewers\n(Warm)", "Watched 15+ sec of\nany Evermore video\nLast 30 days",
     "Traffic / Leads\nCost Cap\n$20 CPL target",
     "$10/day\nRolling 30-day\naudience",
     "Video Ad B\n(15-sec pickup — the PROBLEM)", LORANGE, "7B3F00"),
    ("MOF-2\n50% Video Viewers\n(Warm)", "Watched 50%+ of\nany Evermore video\nLast 30 days",
     "Traffic / Leads\nCost Cap\n$18 CPL target",
     "$10/day\nRolling 30-day",
     "Video Ad C\n(50% pickup — the SOLUTION)", LORANGE, "7B3F00"),
    ("MOF-3\n75-90% Viewers\n(Hot-ish)", "Watched 75%+ of\nany Evermore video\nLast 14 days",
     "Leads / Conversions\nCost Cap\n$15 CPL target",
     "$10/day\nRolling 14-day",
     "Video Ad D\n(75% pickup — social proof)", LGREEN, "155724"),
    ("MOF-4\nWebsite Visitors\n(Warm)", "Visited Sarah's page\nor landing page\nLast 14 days",
     "Leads / Conversions\nCost Cap\n$15 CPL",
     "$8/day\nRolling 14-day",
     "Image Ad C\n('You were this close...')", LGREEN, "155724"),

    # BOTTOM OF FUNNEL
    ("── BOTTOM OF FUNNEL ──", "", "", "", "", RED, WHITE),
    ("BOF-1\n100% Viewers +\nEngagers (Hot)", "Watched 100% of video\nOR page engagers\nLast 14 days",
     "Leads / Conversions\nCost Cap\n$12 CPL",
     "$10/day\nRolling 14-day",
     "Video Ad E\n(Direct ask — book a call)", LRED, RED),
    ("BOF-2\nLead Form\nAbandoners", "Opened lead form\nbut did NOT submit\nLast 7 days",
     "Leads\nCost Cap\n$10 CPL",
     "$5/day\nRolling 7-day",
     "Image Ad D\n('We saved your spot')", LRED, RED),
    ("BOF-3\nSarah Chat\nAbandoners", "Visited Sarah URL\n+ scrolled or clicked\nbut no booking\nLast 7 days",
     "Conversions\nCost Cap\n$10 CPL",
     "$5/day\nRolling 7-day",
     "Image Ad E\n(Sarah follow-up / urgency)", LRED, RED),

    # LOOKALIKE
    ("── SCALE: LOOKALIKE AUDIENCES ──", "", "", "", "", "155724", WHITE),
    ("SCALE-1\n1% Lookalike\n(New Cold)", "1% LAL of\nlead form submitters\n(need 100+ to build)",
     "Lead Generation\nLowest Cost",
     "$20/day\nOnce LAL is live",
     "Best-performing\ncreative from TOF", LGREEN, "155724"),
    ("SCALE-2\n1% Lookalike\nPolicy Closers", "1% LAL of\nactual policy closers\n(highest quality signal)",
     "Lead Generation\nLowest Cost",
     "$25/day\nHighest priority",
     "Same as SCALE-1\nthen iterate", LGREEN, "155724"),
]

current_row = 5
for row_data in ad_sets:
    if len(row_data) == 7:
        name, aud, obj, budget, creative, bg, fg = row_data
        is_section = aud == ""
    else:
        continue

    if is_section:
        ws3.merge_cells(f"B{current_row}:F{current_row}")
        ws3[f"B{current_row}"].value = name
        ws3[f"B{current_row}"].fill = fill(bg)
        ws3[f"B{current_row}"].font = Font(bold=True, size=11, color=fg, name="Arial")
        ws3[f"B{current_row}"].alignment = align("center", "center")
        ws3.row_dimensions[current_row].height = 22
    else:
        for i, (val, col) in enumerate([(name, "B"), (aud, "C"), (obj, "D"), (budget, "E"), (creative, "F")]):
            ws3[f"{col}{current_row}"].value = val
            ws3[f"{col}{current_row}"].fill = fill(bg)
            ws3[f"{col}{current_row}"].font = Font(
                bold=(i == 0), size=9 if i > 0 else 10,
                color=fg if i == 0 else "333333", name="Arial"
            )
            ws3[f"{col}{current_row}"].alignment = align("left", "top", wrap=True)
            ws3[f"{col}{current_row}"].border = border()
        ws3.row_dimensions[current_row].height = 52
    current_row += 1

freeze(ws3, "B5")


# ══════════════════════════════════════════════════════════════
#  SHEET 4: AD SCRIPTS & COPY
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Ad Scripts & Copy")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = GREEN

set_col_width(ws4, {"A":2,"B":22,"C":70,"D":2})

ws4.merge_cells("B2:C2")
ws4["B2"].value = "EVERMORE LIFE — AD SCRIPTS, COPY & CTAs"
ws4["B2"].fill = fill(NAVY)
ws4["B2"].font = Font(bold=True, size=14, color=GOLD, name="Arial")
ws4["B2"].alignment = align("center", "center")
ws4.row_dimensions[2].height = 32

ads = [
    ("VIDEO AD A — HERO / TOF (60–90 sec)", "FULL STORY AD — runs cold to new audiences. This is the foundation of the whole funnel.", [
        ("Hook (0–5 sec)", "\"If something happened to you this week, who would pay for your funeral?\"\n[Pause. Let it land.]\n\"I'm not trying to scare you — I just need you to hear this.\""),
        ("Problem (5–20 sec)", "\"The average funeral today costs between $12,000 and $15,000. And your family would have to come up with that money — fast. Within days. While they're grieving.\"\n\"Most people think their life insurance or Social Security will cover it. Most of the time... it doesn't. Not enough.\""),
        ("Introduce Solution (20–40 sec)", "\"That's why final expense insurance exists. It's a small, affordable policy — specifically designed to cover end-of-life costs.\"\n\"No medical exam for most plans. Your rate is locked in for life — it never goes up. And coverage starts immediately for accidental death.\"\n\"Plans start as low as $1 a day for many people. Sometimes less.\""),
        ("Credibility (40–55 sec)", "\"At Evermore Life, we're licensed in your state. We don't work for one company — we shop dozens of carriers to find the lowest rate your health qualifies you for.\"\n\"We do the work. You make one call.\""),
        ("CTA (55–90 sec)", "\"If you're between 50 and 85, click the button below. Answer a few quick questions — no obligation, completely free. We'll show you exactly what you qualify for.\"\n\"Protect your family. Leave peace behind.\"\n[On screen: EVERMORE LIFE logo + 'Get Your Free Quote']"),
        ("Caption / Primary Text", "⚠️ Your family could face $15,000+ in burial costs.\n\nFinal expense insurance can protect them — starting under $1/day for many ages.\n\n✅ No medical exam required for most plans\n✅ Rate locked in for life — never increases\n✅ Licensed agents in your state\n✅ We shop all the top carriers for you\n\nFind out what you qualify for — free, no obligation. 👇"),
        ("Headline", "\"Protect Your Family From Burial Debt\""),
        ("CTA Button", "Get Free Quote"),
    ]),
    ("VIDEO AD B — 15-SEC RETARGETING (PROBLEM)", "Served to: 15-second video viewers. They saw the hook. Now deepen the problem.", [
        ("Hook (0–3 sec)", "\"You saw this video earlier. I want to make sure you really understood what I was saying.\""),
        ("Deepen Problem (3–18 sec)", "\"If you pass away without final expense coverage, your family has about 72 hours to come up with thousands of dollars. Some families go into debt. Some can't afford a proper burial. Some never fully recover financially.\"\n\"You've worked your whole life to take care of them. Don't leave them with this.\""),
        ("Soft CTA (18–30 sec)", "\"It takes less than 2 minutes to find out what you qualify for. Click below.\""),
        ("Caption / Primary Text", "You started watching our video. Did this hit home?\n\nMost people assume someone else will handle it. Don't let your family find out the hard way.\n\nFind out your rate — free, no obligation. Takes 2 minutes. 👇"),
        ("Headline", "\"Don't Leave Your Family With This Burden\""),
        ("CTA Button", "See If I Qualify"),
    ]),
    ("VIDEO AD C — 50% RETARGETING (SOLUTION)", "Served to: 50% video viewers. They understand the problem. Now give them the solution.", [
        ("Hook (0–3 sec)", "\"Okay — you understand the problem. Let me show you exactly how final expense insurance works.\""),
        ("Mechanism (3–20 sec)", "\"You apply — we ask a few simple health questions. No doctor visit. No blood test. Based on your age and health, we find the best rate from our network of carriers.\"\n\"You pay a small monthly premium — usually between $30 and $90 for most ages. It never goes up. And the benefit goes directly to your family — $10,000, $15,000, $20,000 — to handle everything.\""),
        ("Handle objections (20–35 sec)", "\"'What if I have health issues?' — Most conditions still qualify. We have guaranteed-issue plans for harder-to-insure situations.\"\n\"'What if I already have life insurance?' — Final expense is different. It's designed specifically for this moment. Many people have both.\""),
        ("CTA (35–45 sec)", "\"Click below. Talk to Sarah — our AI assistant — or go straight to booking a free call with one of our licensed agents. Either way, there's no pressure and no cost.\""),
        ("Caption / Primary Text", "You know the problem. Here's the solution.\n\nFinal expense insurance = small monthly payment → your family gets a lump sum when you pass.\n\n🔒 Rate locked for life\n📋 No medical exam for most\n💰 Starts under $1/day\n🏆 We shop all carriers for your best rate\n\nSee your options — free. 👇"),
        ("Headline", "\"Here's How Final Expense Insurance Works\""),
        ("CTA Button", "See My Options"),
    ]),
    ("VIDEO AD D — 75% RETARGETING (SOCIAL PROOF)", "Served to: 75–90% video viewers. Nearly convinced. Use proof and urgency.", [
        ("Hook (0–3 sec)", "\"You've done the research. You know this makes sense. Let me give you one more reason to act today.\""),
        ("Social Proof (3–20 sec)", "\"Every week, families call us after losing a loved one — wishing they had done this sooner. And every week, we help people just like you get covered in one conversation.\"\n\"[Optional: brief agent quote or testimonial graphic]\"\n\"At Evermore Life — simple, honest, trusted.\""),
        ("Urgency (20–30 sec)", "\"Here's the thing: the older you get, the higher your rate. Every year you wait costs you money. The best time to lock in your rate is right now.\"\n\"Don't put this off. Your family is counting on you.\""),
        ("CTA (30–40 sec)", "\"Click below. Book a free 15-minute call. Our licensed agents are ready.\"\n\"No sales pressure. Just honest answers.\""),
        ("Caption / Primary Text", "You've been thinking about this. Let's get it done today.\n\nEvery year you wait = higher rates. Lock yours in now.\n\n\"I wish we'd done this sooner\" — something families say too often.\n\nBook a free 15-min call with a licensed Evermore Life agent. No pressure. Just answers. 👇"),
        ("Headline", "\"Lock In Your Rate Before It Goes Up\""),
        ("CTA Button", "Book Free Consultation"),
    ]),
    ("VIDEO AD E — 100% VIEWERS / BOTTOM OF FUNNEL (DIRECT ASK)", "Served to: 100% viewers + page engagers. They know everything. Make the direct ask.", [
        ("Hook (0–3 sec)", "\"You've watched everything. You know what this is. You know why it matters. The only thing left is to make a decision.\""),
        ("Reframe (3–15 sec)", "\"Here's what happens next: you click below, book a free 15-minute call, and our licensed agent walks you through exactly what you qualify for — rates, coverage amounts, carriers. Zero pressure.\"\n\"If it's right for you, great. If not, no hard feelings. But at least you'll know.\""),
        ("Close (15–25 sec)", "\"Most people who get on this call say they wish they'd done it sooner.\"\n\"Protect what matters. Leave peace behind.\"\n\"— Evermore Life\""),
        ("Caption / Primary Text", "You've done the research. Now let's get you covered.\n\nBook a FREE 15-minute consultation with a licensed Evermore Life agent.\n\n🗓️ Pick a time that works for you\n📞 One quick call — we handle everything\n✅ No obligation, no pressure\n\nSpots are limited. Click below before yours is gone. 👇"),
        ("Headline", "\"Ready? Let's Get You Covered.\""),
        ("CTA Button", "Book My Free Call"),
    ]),
    ("IMAGE AD A — TOF LEAD FORM", "Static image. Paired with lead form. Simple, direct.", [
        ("Visual Description", "Background: Warm family photo (grandparent with grandchildren, golden hour lighting). Overlay: Evermore Life logo top-left. Bold text center: 'FREE FINAL EXPENSE QUOTE'\nSubtext: 'No exam. No obligation. Rates from $X/mo*'\nBottom bar: navy with gold text 'Evermore Life | Licensed in [Your State]'"),
        ("Primary Text", "Worried about leaving your family with funeral costs?\n\nFinal expense insurance can cover $10,000–$25,000 in end-of-life expenses — starting under $1/day for many ages.\n\n✅ No medical exam for most plans\n✅ Rate locked for life\n✅ Licensed agents in your state\n\nGet your free quote in 2 minutes. No obligation. 👇"),
        ("Headline", "Get Your Free Final Expense Quote Today"),
        ("Description", "Licensed agents. Multiple carriers. Find your best rate."),
        ("CTA Button", "Get Quote"),
    ]),
    ("IMAGE AD B — CHILDREN OF SENIORS ANGLE", "Target: adult children (35–55) researching coverage for a parent.", [
        ("Visual Description", "Image: Adult child (40s) sitting with elderly parent at kitchen table, warm/cozy. Text overlay: 'Protect Mom or Dad Before It's Too Late'\nEvermore Life logo. Navy/gold color scheme."),
        ("Primary Text", "Have you thought about what happens to your parent's final expenses?\n\nThe average funeral costs $12,000–$15,000. And that bill comes due within days.\n\nHelp them get covered NOW — before their rates go up.\n\nEvermore Life helps seniors get final expense coverage:\n✅ No medical exam for most\n✅ Affordable monthly payments\n✅ We shop ALL the top carriers\n\nGet a free quote for your parent in 2 minutes. 👇"),
        ("Headline", "\"Give Mom or Dad the Gift of Peace of Mind\""),
        ("CTA Button", "Get Free Quote"),
    ]),
    ("IMAGE AD C — WEBSITE RETARGETING ('You Were This Close')", "Served to: people who visited Sarah's page but didn't convert.", [
        ("Visual Description", "Bold text: 'You Were This Close.'\nSubtext: 'Your family's protection was one click away.'\nEvermore Life logo. Dark navy background, gold accent.\nSmall: 'Still free. Still no obligation. Still here.'"),
        ("Primary Text", "You checked us out. Something stopped you.\n\nMaybe it wasn't the right time. Maybe you had questions.\n\nWe get it. No pressure.\n\nBut here's the truth: the longer you wait, the more your rates go up. Every. Single. Year.\n\nCome back. It takes 2 minutes. We'll answer every question you have — free. 👇"),
        ("Headline", "\"We're Still Here When You're Ready\""),
        ("CTA Button", "Continue Where I Left Off"),
    ]),
    ("IMAGE AD D — LEAD FORM ABANDONMENT", "Served to: people who opened the lead form but didn't submit.", [
        ("Visual Description", "Text: 'We Saved Your Spot.'\nSubtext: 'Your free quote is waiting — takes 60 seconds to finish.'\nProgress bar graphic showing '80% Complete'\nEvermore Life branding."),
        ("Primary Text", "You started your free quote. Life got in the way — we get it.\n\nYour information is safe. Your spot is saved.\n\nJust finish your quote — it takes less than 60 seconds.\n\nWe'll match you with the lowest rate you qualify for. No exam. No obligation. 👇"),
        ("Headline", "\"Finish Your Free Quote (60 Seconds)\""),
        ("CTA Button", "Finish My Quote"),
    ]),
    ("IMAGE AD E — SARAH FOLLOW-UP / URGENCY", "Served to: Sarah chat abandoners. They talked to Sarah but didn't book.", [
        ("Visual Description", "Sarah's office visual (pulled from the website). Text: 'Sarah is still waiting for you.'\nSubtext: 'Your appointment slot is reserved for 24 hours.'\nEvermore Life branding. Warm, not pushy."),
        ("Primary Text", "Hey — Sarah noticed you didn't finish booking your consultation.\n\nYour spot is reserved for the next 24 hours.\n\nAll it takes is 15 minutes on the phone with a real licensed agent. They'll answer every question Sarah couldn't — and show you exactly what you qualify for.\n\nNo pressure. Just honest answers. From real people. 👇"),
        ("Headline", "\"Your Consultation Slot Is Waiting\""),
        ("CTA Button", "Book My Spot Now"),
    ]),
]

current_row = 5
for ad_title, ad_desc, elements in ads:
    # Ad header
    ws4.merge_cells(f"B{current_row}:C{current_row}")
    ws4[f"B{current_row}"].value = ad_title
    ws4[f"B{current_row}"].fill = fill(NAVY)
    ws4[f"B{current_row}"].font = Font(bold=True, size=11, color=GOLD, name="Arial")
    ws4[f"B{current_row}"].alignment = align("left", "center")
    ws4.row_dimensions[current_row].height = 24
    current_row += 1

    # Ad description
    ws4.merge_cells(f"B{current_row}:C{current_row}")
    ws4[f"B{current_row}"].value = ad_desc
    ws4[f"B{current_row}"].fill = fill(LGOLD)
    ws4[f"B{current_row}"].font = Font(italic=True, size=9, color="555555", name="Arial")
    ws4[f"B{current_row}"].alignment = align("left", "center")
    ws4.row_dimensions[current_row].height = 18
    current_row += 1

    for label, content in elements:
        ws4[f"B{current_row}"].value = label
        ws4[f"B{current_row}"].fill = fill(LNAVY)
        ws4[f"B{current_row}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
        ws4[f"B{current_row}"].alignment = align("left", "top", wrap=True)
        ws4[f"B{current_row}"].border = border()

        ws4[f"C{current_row}"].value = content
        ws4[f"C{current_row}"].fill = fill(WHITE)
        ws4[f"C{current_row}"].font = Font(size=10, color="222222", name="Arial")
        ws4[f"C{current_row}"].alignment = align("left", "top", wrap=True)
        ws4[f"C{current_row}"].border = border()

        lines = content.count("\n") + 1
        ws4.row_dimensions[current_row].height = max(30, min(lines * 15, 120))
        current_row += 1

    current_row += 1

freeze(ws4, "B3")


# ══════════════════════════════════════════════════════════════
#  SHEET 5: IMAGE PROMPTS
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Image Prompts (AI)")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = ORANGE

set_col_width(ws5, {"A":2,"B":22,"C":26,"D":55,"E":2})

ws5.merge_cells("B2:D2")
ws5["B2"].value = "EVERMORE LIFE — AI IMAGE GENERATION PROMPTS"
ws5["B2"].fill = fill(NAVY)
ws5["B2"].font = Font(bold=True, size=14, color=GOLD, name="Arial")
ws5["B2"].alignment = align("center", "center")
ws5.row_dimensions[2].height = 32

ws5.merge_cells("B3:D3")
ws5["B3"].value = "Use these prompts in Midjourney, DALL-E 3, or Adobe Firefly. Adjust text overlays in Canva or Adobe Express."
ws5["B3"].fill = fill(LGOLD)
ws5["B3"].font = Font(italic=True, size=9, color="555555", name="Arial")
ws5["B3"].alignment = align("left", "center")
ws5.row_dimensions[3].height = 18

img_headers = ["AD", "FORMAT / SIZE", "AI IMAGE PROMPT"]
for i, h in enumerate(img_headers):
    col = get_column_letter(i + 2)
    ws5[f"{col}5"].value = h
    ws5[f"{col}5"].fill = fill(NAVY)
    ws5[f"{col}5"].font = Font(bold=True, size=10, color=WHITE, name="Arial")
    ws5[f"{col}5"].alignment = align("center", "center")
    ws5[f"{col}5"].border = border()
ws5.row_dimensions[5].height = 22

image_prompts = [
    ("Video Ad A\n(Thumbnail)", "1280×720 (16:9)\nYouTube/Facebook\nvideo thumbnail",
     "A warm, emotional scene of an elderly couple (70s) sitting on a front porch at golden hour, holding hands, looking peaceful. Soft bokeh background of a green suburban yard. Cinematic lighting. Photorealistic. No text. High resolution. Canon 5D look. Color palette: warm amber and soft navy shadows."),
    ("Image Ad A\n(Lead Form)", "1200×628 (1.91:1)\nFacebook Feed",
     "A multi-generational family (grandparents, adult children, grandchildren) gathered in a warm living room, laughing together. Golden hour light through large windows. Photorealistic. Warm amber tones. Navy blue accent in decor. Cozy and authentic. No logos or text in image. High resolution."),
    ("Image Ad B\n(Children of Seniors)", "1200×628 (1.91:1)\nFacebook Feed",
     "An adult woman in her 40s sitting at a kitchen table with her elderly mother (70s), both looking at papers or a tablet, expressions warm and caring. Kitchen is cozy with warm lighting. Photorealistic. Authentic, not staged. No text. Canon EOS R5 style photography."),
    ("Image Ad C\n(Retargeting — 'Close')", "1200×628 (1.91:1)\nFacebook Feed",
     "A lone empty chair at a simple wooden kitchen table with morning sunlight streaming in. One coffee cup. Quiet, contemplative mood. Slightly melancholic but peaceful. Photorealistic. No people. Symbolizes absence. Muted warm tones with navy shadows. No text."),
    ("Image Ad D\n(Lead Form Abandonment)", "1200×628 (1.91:1)\nFacebook Feed",
     "A smartphone screen showing a simple insurance quote form, partially filled in, resting on a wooden desk next to a cup of coffee and a pen. Overhead flat lay shot. Clean and modern. Soft shadows. No specific brand text. Photorealistic."),
    ("Image Ad E\n(Sarah Follow-Up)", "1200×628 (1.91:1)\nFacebook Feed",
     "A professional, warm office environment with a wooden desk, bookshelf with books, a laptop, a small potted plant, and soft warm lamp light. Empty desk chair. Welcoming atmosphere. The kind of office you'd feel comfortable walking into. No people. Photorealistic. Navy and warm gold accent colors."),
    ("Video Ad B Thumbnail\n(15-sec retargeting)", "1280×720 (16:9)\nFacebook Video",
     "Close-up of an elderly woman's hands resting on a table, next to a handwritten note that reads 'For my family.' Soft focus background of a modest home. Warm amber lighting. Emotional, tender. Photorealistic. No harsh shadows. Canon 5D portrait look."),
    ("Video Ad E Thumbnail\n(BOF — Book a Call)", "1280×720 (16:9)\nFacebook Video",
     "A friendly, professional male insurance agent (50s, business casual) sitting across from an elderly couple at a kitchen table, all smiling and reviewing papers together. Natural home lighting. Warm, trustworthy atmosphere. Photorealistic. No logos or text in image."),
    ("Sarah AI Avatar\n(Profile / Chatbot icon)", "400×400 (1:1)\nProfile / Chat widget",
     "A professional, friendly female AI assistant illustration. Business casual attire (navy blazer). Warm smile. Approachable and trustworthy. Slight stylized illustration style (not fully photorealistic — more like a polished corporate illustration). Clean white or soft gradient background. No text. Navy blue and gold accents. Would work as a profile picture for a chat interface."),
    ("Evermore Life\nFunnel Header Banner", "1640×624 (Facebook\nPage cover ratio)",
     "Wide panoramic banner image: A family (grandparents, parents, children) standing together on a hill at golden hour, looking out at a sunset over a peaceful landscape. Silhouette-style with warm amber and deep navy tones. Cinematic. Emotional. Sense of legacy and protection. No people's faces visible — universal family silhouette. No text. Photorealistic cinematic style."),
]

for i, (ad_name, fmt, prompt) in enumerate(image_prompts):
    row = i + 6
    ws5[f"B{row}"].value = ad_name
    ws5[f"B{row}"].fill = fill(LGOLD if i % 2 == 0 else LORANGE)
    ws5[f"B{row}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
    ws5[f"B{row}"].alignment = align("left", "top", wrap=True)
    ws5[f"B{row}"].border = border()

    ws5[f"C{row}"].value = fmt
    ws5[f"C{row}"].fill = fill(LNAVY)
    ws5[f"C{row}"].font = Font(size=9, color="333333", name="Arial")
    ws5[f"C{row}"].alignment = align("center", "top", wrap=True)
    ws5[f"C{row}"].border = border()

    ws5[f"D{row}"].value = prompt
    ws5[f"D{row}"].fill = fill(WHITE)
    ws5[f"D{row}"].font = Font(size=10, color="111111", name="Arial")
    ws5[f"D{row}"].alignment = align("left", "top", wrap=True)
    ws5[f"D{row}"].border = border()

    lines = prompt.count(".") + 1
    ws5.row_dimensions[row].height = max(45, min(lines * 18, 100))

freeze(ws5, "B6")


# ══════════════════════════════════════════════════════════════
#  SHEET 6: METRICS & KPIs
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. KPIs & Budget")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = RED

set_col_width(ws6, {"A":3,"B":26,"C":20,"D":20,"E":20,"F":3})

ws6.merge_cells("B2:E2")
ws6["B2"].value = "EVERMORE LIFE — KPIs, BENCHMARKS & BUDGET TRACKER"
ws6["B2"].fill = fill(NAVY)
ws6["B2"].font = Font(bold=True, size=14, color=GOLD, name="Arial")
ws6["B2"].alignment = align("center", "center")
ws6.row_dimensions[2].height = 32

# KPI table
kpi_headers = ["METRIC", "TARGET", "WARNING", "PAUSE / FIX"]
kpi_row = 4
for i, h in enumerate(kpi_headers):
    col = get_column_letter(i + 2)
    ws6[f"{col}{kpi_row}"].value = h
    ws6[f"{col}{kpi_row}"].fill = fill(NAVY)
    ws6[f"{col}{kpi_row}"].font = Font(bold=True, size=10, color=WHITE, name="Arial")
    ws6[f"{col}{kpi_row}"].alignment = align("center", "center")
    ws6[f"{col}{kpi_row}"].border = border()
ws6.row_dimensions[kpi_row].height = 22

kpis = [
    ("Cost Per Lead (CPL)",             "< $25",        "$25–$40",      "> $40 — pause, test new creative"),
    ("Cost Per View (Video — 15-sec)",  "< $0.05",      "$0.05–$0.10",  "> $0.10 — test new hook (first 3 sec)"),
    ("Lead Form Completion Rate",       "> 15%",        "10–15%",       "< 10% — simplify form, test new headline"),
    ("Video View Rate (3-sec)",         "> 25%",        "15–25%",       "< 15% — kill the ad, new hook"),
    ("ThruPlay Rate (full view)",       "> 8%",         "4–8%",         "< 4% — ad is too long or loses attention"),
    ("Sarah Booking Rate",              "> 20% of chats","10–20%",      "< 10% — improve Sarah conversation flow"),
    ("Appointment Show Rate",           "> 40%",        "25–40%",       "< 25% — add SMS reminder, pre-call sequence"),
    ("Lead → Policy Close Rate",        "> 20%",        "10–20%",       "< 10% — review agent pitch, lead quality"),
    ("Retargeting CPL vs TOF CPL",      "50% lower",    "25–50% lower", "If equal — audiences too small, expand TOF"),
    ("Frequency (TOF cold audiences)",  "1.5–3x / 7d",  "3–5x / 7d",   "> 5x — expand audience or refresh creative"),
]

for i, (metric, target, warn, pause) in enumerate(kpis):
    r = kpi_row + 1 + i
    bg = LGRAY if i % 2 == 0 else WHITE
    ws6[f"B{r}"].value = metric
    ws6[f"B{r}"].fill = fill(bg)
    ws6[f"B{r}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
    ws6[f"B{r}"].border = border()
    ws6[f"B{r}"].alignment = align("left", "center")

    ws6[f"C{r}"].value = target
    ws6[f"C{r}"].fill = fill(LGREEN)
    ws6[f"C{r}"].font = Font(bold=True, size=10, color=GREEN, name="Arial")
    ws6[f"C{r}"].border = border()
    ws6[f"C{r}"].alignment = align("center", "center")

    ws6[f"D{r}"].value = warn
    ws6[f"D{r}"].fill = fill(LORANGE)
    ws6[f"D{r}"].font = Font(size=10, color=ORANGE, name="Arial")
    ws6[f"D{r}"].border = border()
    ws6[f"D{r}"].alignment = align("center", "center", wrap=True)

    ws6[f"E{r}"].value = pause
    ws6[f"E{r}"].fill = fill(LRED)
    ws6[f"E{r}"].font = Font(size=9, color=RED, name="Arial")
    ws6[f"E{r}"].border = border()
    ws6[f"E{r}"].alignment = align("left", "center", wrap=True)

    ws6.row_dimensions[r].height = 28

# Budget breakdown
budget_start = kpi_row + len(kpis) + 3
ws6.merge_cells(f"B{budget_start}:E{budget_start}")
ws6[f"B{budget_start}"].value = "DAILY BUDGET ALLOCATION — STARTER ($50/day total)"
ws6[f"B{budget_start}"].fill = fill(NAVY)
ws6[f"B{budget_start}"].font = Font(bold=True, size=11, color=GOLD, name="Arial")
ws6[f"B{budget_start}"].alignment = align("center", "center")
ws6.row_dimensions[budget_start].height = 24

budget_row = budget_start + 1
budget_headers = ["AD SET", "DAILY BUDGET", "% OF TOTAL", "NOTES"]
for i, h in enumerate(budget_headers):
    col = get_column_letter(i + 2)
    ws6[f"{col}{budget_row}"].value = h
    ws6[f"{col}{budget_row}"].fill = fill(LNAVY)
    ws6[f"{col}{budget_row}"].font = Font(bold=True, size=10, color=NAVY, name="Arial")
    ws6[f"{col}{budget_row}"].border = border()
    ws6[f"{col}{budget_row}"].alignment = align("center", "center")
ws6.row_dimensions[budget_row].height = 22

budget_data = [
    ("TOF-1 Broad 50–75",            "$15",  "30%", "Primary video views builder"),
    ("TOF-2 Interest Stack",          "$15",  "30%", "Primary lead gen"),
    ("MOF-1 15-sec retargeting",      "$5",   "10%", "Reactivate early dropoffs"),
    ("MOF-2 50% viewers",             "$5",   "10%", "Solution messaging"),
    ("BOF-1 100% viewers / engagers", "$5",   "10%", "Direct close attempt"),
    ("BOF-2 Form abandoners",         "$3",   "6%",  "Quick re-engage"),
    ("BOF-3 Sarah abandoners",        "$2",   "4%",  "Booking recovery"),
    ("TOTAL",                         "$50",  "100%","Scale winners 2x every 5 days"),
]

for i, (name, bud, pct, notes) in enumerate(budget_data):
    r = budget_row + 1 + i
    is_total = name == "TOTAL"
    bg = NAVY if is_total else (LGRAY if i % 2 == 0 else WHITE)
    fg_color = GOLD if is_total else "333333"

    ws6[f"B{r}"].value = name
    ws6[f"B{r}"].fill = fill(bg)
    ws6[f"B{r}"].font = Font(bold=is_total, size=10, color=fg_color, name="Arial")
    ws6[f"B{r}"].border = border()

    ws6[f"C{r}"].value = bud
    ws6[f"C{r}"].fill = fill(bg)
    ws6[f"C{r}"].font = Font(bold=is_total, size=10, color=fg_color, name="Arial")
    ws6[f"C{r}"].border = border()
    ws6[f"C{r}"].alignment = align("center", "center")

    ws6[f"D{r}"].value = pct
    ws6[f"D{r}"].fill = fill(bg)
    ws6[f"D{r}"].font = Font(bold=is_total, size=10, color=fg_color, name="Arial")
    ws6[f"D{r}"].border = border()
    ws6[f"D{r}"].alignment = align("center", "center")

    ws6[f"E{r}"].value = notes
    ws6[f"E{r}"].fill = fill(bg)
    ws6[f"E{r}"].font = Font(italic=(not is_total), size=9, color=fg_color, name="Arial")
    ws6[f"E{r}"].border = border()
    ws6[f"E{r}"].alignment = align("left", "center", wrap=True)

    ws6.row_dimensions[r].height = 22

freeze(ws6, "B4")


# ── Save ──────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[2]
out = PROJECT_ROOT / "03_sales_marketing" / "playbooks" / "Evermore_Meta_Funnel_Playbook.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out))
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
