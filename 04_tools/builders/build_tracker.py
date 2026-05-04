from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = Workbook()
ws = wb.active
ws.title = "Implementation Tracker"
ws.sheet_view.showGridLines = False

NAVY   = "0D1B3E"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
GREEN  = "1E7E34"
LGREEN = "D4EDDA"
RED    = "C0392B"
LRED   = "FADBD8"
ORANGE = "D4720A"
LORANGE= "FFF3CD"
LGRAY  = "F8F9FA"
LNAVY  = "E8ECF5"

def fill(c): return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 5
ws.column_dimensions["C"].width = 36
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 48
ws.column_dimensions["F"].width = 3

# Title
ws.merge_cells("B2:E2")
ws["B2"].value = "EVERMORE LIFE — MASTER IMPLEMENTATION TRACKER"
ws["B2"].fill = fill(NAVY)
ws["B2"].font = Font(bold=True, size=16, color=GOLD, name="Arial")
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 36

ws.merge_cells("B3:E3")
ws["B3"].value = "🟢 DONE (built & ready)   |   🟡 YOU DO IT (requires your account/action)   |   🔵 IN PROGRESS (building now)"
ws["B3"].fill = fill(LNAVY)
ws["B3"].font = Font(italic=True, size=10, color="333333", name="Arial")
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 20

# Header row
for col, txt in [(2,"#"),(3,"TASK / DELIVERABLE"),(4,"STATUS"),(5,"WHAT TO DO / FILE LOCATION")]:
    ws.cell(row=5, column=col).value = txt
    ws.cell(row=5, column=col).fill = fill(NAVY)
    ws.cell(row=5, column=col).font = Font(bold=True, size=10, color=WHITE, name="Arial")
    ws.cell(row=5, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=col).border = bdr()
ws.row_dimensions[5].height = 24

STATUS_DONE     = ("✅ DONE",      LGREEN,  GREEN)
STATUS_YOU      = ("🟡 YOU DO IT", LORANGE, ORANGE)
STATUS_PROGRESS = ("🔵 BUILDING",  LNAVY,   NAVY)
STATUS_FILE     = ("📁 FILE READY",LGREEN,  GREEN)

tasks = [
    # Section, items
    ("SARAH AI ASSISTANT — FRONT-END", [
        ("1",  "Sarah v1 — Virtual office, chat, booking, confirmation", STATUS_FILE,
         "File: Sarah_Evermore_AI.html\nOpen in browser to test. Works as-is."),
        ("2",  "Sarah v2 — Enhanced (8-step flow, UTM tracking, state dropdown, personalized responses)", STATUS_PROGRESS,
         "File: Sarah_Evermore_AI_v2.html\nBuilding now — will be in outputs folder."),
        ("3",  "Meta Pixel events wired into Sarah (PageView, Lead, Schedule)", STATUS_PROGRESS,
         "In Sarah v2 — integration points clearly marked.\nYou add: your Meta Pixel ID where marked."),
        ("4",  "GHL Webhook fired on booking submit", STATUS_PROGRESS,
         "In Sarah v2 — integration point marked.\nYou add: paste your GHL webhook URL into the marked line."),
        ("5",  "Claude API wired for real AI responses", STATUS_YOU,
         "YOU DO IT:\n1. Go to console.anthropic.com\n2. Create API key\n3. In Sarah v2 — find INTEGRATION POINT 1\n4. Uncomment the fetch() call and paste your API key\nSystem prompt is already written for you."),
        ("6",  "Deploy Sarah to live URL", STATUS_YOU,
         "YOU DO IT (on home Mac):\nOption A: Cloudflare Workers — copy deploy_kvn.sh pattern, swap KVN.html for Sarah_Evermore_AI_v2.html\nOption B: Paste into a GHL funnel page as custom HTML block\nOption C: Vercel.com — drag & drop, free"),
    ]),
    ("KVN — INTERNAL AGENT", [
        ("7",  "KVN.html — Full 3D character, boot screen, chat, mobile tabs", STATUS_FILE,
         "File: KVN.html\nAll features built. Ready to deploy."),
        ("8",  "KVN deploy script (Cloudflare Workers)", STATUS_FILE,
         "File: deploy_kvn.sh\nRun this in Terminal on your home Mac. Takes 10 seconds.\nURL: https://kvn-agent.newmexicomarketingpartners.workers.dev"),
        ("9",  "KVN — Deploy to live URL", STATUS_YOU,
         "YOU DO IT (home Mac):\nOpen Terminal → run deploy_kvn.sh\nVerify URL loads on your phone."),
        ("10", "KVN — Wire to Claude API for real intelligence", STATUS_YOU,
         "YOU DO IT:\n1. Get API key from console.anthropic.com\n2. In KVN.html find INTEGRATION POINT\n3. Paste key + uncomment the API call\nKVN system prompt: already written in the file."),
        ("11", "KVN — Daily briefing mode", STATUS_YOU,
         "Future build — needs Google Calendar API or manual schedule input.\nTell Claude when you're ready to build this feature."),
    ]),
    ("META ADS — CONTENT & COPY", [
        ("12", "Video Ad A script (Hero / 60-90 sec)", STATUS_PROGRESS,
         "File: Evermore_Video_Ad_Scripts.docx\nFull teleprompter-ready script. Building now."),
        ("13", "Video Ad B script (15-sec retargeting)", STATUS_PROGRESS,
         "Same file: Evermore_Video_Ad_Scripts.docx"),
        ("14", "Video Ad C script (50% retargeting)", STATUS_PROGRESS,
         "Same file: Evermore_Video_Ad_Scripts.docx"),
        ("15", "Video Ad D script (75% retargeting)", STATUS_PROGRESS,
         "Same file: Evermore_Video_Ad_Scripts.docx"),
        ("16", "Video Ad E script (BOF — direct ask)", STATUS_PROGRESS,
         "Same file: Evermore_Video_Ad_Scripts.docx"),
        ("17", "All image ad primary text + headlines + CTAs", STATUS_FILE,
         "File: Evermore_Meta_Funnel_Playbook.xlsx\nTab 4: Ad Scripts & Copy — all 10 ads fully written."),
        ("18", "AI image generation prompts (all 10 ads)", STATUS_FILE,
         "File: Evermore_Meta_Funnel_Playbook.xlsx\nTab 5: Image Prompts — ready to paste into Midjourney/DALL-E."),
        ("19", "Record Video Ad A (the hero 60-90 sec video)", STATUS_YOU,
         "YOU DO IT:\nUse the script from Evermore_Video_Ad_Scripts.docx\nOptions: Record yourself on iPhone | Hire UGC creator on Billo.app or Fiverr\nEdit in CapCut (free) or hire editor\nCaptions: auto-generate in CapCut"),
        ("20", "Generate ad images using AI prompts", STATUS_YOU,
         "YOU DO IT:\n1. Open Evermore_Meta_Funnel_Playbook.xlsx → Tab 5\n2. Copy each prompt\n3. Paste into: Midjourney (best), DALL-E 3 (ChatGPT Plus), or Adobe Firefly (free)\n4. Add text overlays in Canva (free)"),
        ("21", "Record retargeting Videos B-E", STATUS_YOU,
         "YOU DO IT:\nAfter Video A is proven (getting views/leads), record B-E\nShorter = easier. B and E are under 30 seconds."),
    ]),
    ("META ADS — CAMPAIGN SETUP", [
        ("22", "Meta Business Manager account", STATUS_YOU,
         "YOU DO IT:\nbusiness.facebook.com → Create account\nVerify business, add payment method\nConnect your Facebook Page for Evermore Life"),
        ("23", "Meta Pixel installed on Sarah's page", STATUS_YOU,
         "YOU DO IT:\n1. In Meta Events Manager → Create Pixel → Copy Pixel ID\n2. In Sarah_Evermore_AI_v2.html → find META PIXEL comment\n3. Replace YOUR_PIXEL_ID with your actual ID\n4. Deploy updated file\nPixel events already coded: PageView, Lead, Schedule"),
        ("24", "Campaign 1: Video Views (TOF)", STATUS_YOU,
         "YOU DO IT (in Meta Ads Manager):\nObjective: Video Views\n3 ad sets: TOF-1 Broad, TOF-2 Interest Stack, TOF-3 Children of Seniors\n(Full setup in Evermore_Meta_Funnel_Playbook.xlsx Tab 3)\nBudget: $15/day each"),
        ("25", "Campaign 2: Lead Generation (TOF)", STATUS_YOU,
         "YOU DO IT:\nObjective: Leads\nConnect GHL lead form or use native Facebook lead form\nBudget: $15/day\nAudience: same as TOF-2"),
        ("26", "Campaign 3: Traffic (to Sarah page)", STATUS_YOU,
         "YOU DO IT:\nObjective: Traffic\nDestination: Sarah's live URL\nBudget: $10/day\nOnly after Sarah is deployed to live URL"),
        ("27", "Retargeting campaigns (MOF + BOF)", STATUS_YOU,
         "YOU DO IT (after TOF runs 7+ days and builds audiences):\nCreate custom audiences from video viewers (15s, 50%, 75%, 100%)\nCreate ad sets MOF-1 through BOF-3\n(Full details: Evermore_Meta_Funnel_Playbook.xlsx Tab 3)"),
        ("28", "Lookalike audiences (SCALE)", STATUS_YOU,
         "YOU DO IT (after 100+ leads collected):\nCreate 1% LAL from lead form submitters\nCreate 1% LAL from policy closers (if available in GHL)\nNew cold ad sets targeting LAL"),
    ]),
    ("GOHIGHLEVEL CRM", [
        ("29", "GHL account created", STATUS_YOU,
         "YOU DO IT:\ngohighlevel.com → Start 14-day free trial → Pro plan ($297/mo)\nIncludes: CRM, calendar, SMS, email, pipeline, automations"),
        ("30", "GHL pipeline: Final Expense Leads (9 stages)", STATUS_YOU,
         "YOU DO IT (after account setup):\nFollow GHL_Setup_Guide.docx — Section 2\nStages are already defined for you."),
        ("31", "GHL custom contact fields set up", STATUS_YOU,
         "YOU DO IT:\nFollow GHL_Setup_Guide.docx — Section 3\nFields: topic, age, health, coverage, beneficiary, smoker, state, UTMs"),
        ("32", "GHL inbound webhook created", STATUS_YOU,
         "YOU DO IT:\nFollow GHL_Setup_Guide.docx — Section 4\nCopy the webhook URL → paste into Sarah v2 at INTEGRATION POINT 2"),
        ("33", "GHL automation: New Lead Flow (SMS + notifications)", STATUS_YOU,
         "YOU DO IT:\nFollow GHL_Setup_Guide.docx — Section 5\nSMS templates: GHL_SMS_Email_Templates.docx — ready to copy-paste"),
        ("34", "GHL Setup Guide (step-by-step)", STATUS_PROGRESS,
         "File: GHL_Setup_Guide.docx\nBuilding now — detailed walkthrough of every GHL step."),
        ("35", "GHL SMS + Email templates", STATUS_PROGRESS,
         "File: GHL_SMS_Email_Templates.docx\nBuilding now — 6 SMS + 3 email templates, copy-paste ready."),
        ("36", "GHL calendar: Free Consultation (15 min)", STATUS_YOU,
         "YOU DO IT:\nFollow GHL_Setup_Guide.docx — Section 6\nSet your availability, buffer time, confirmation email auto-sends"),
        ("37", "GHL + Meta Pixel Conversions API connected", STATUS_YOU,
         "YOU DO IT:\nFollow GHL_Setup_Guide.docx — Section 7\nIn GHL: Settings → Integrations → Facebook → Connect\nEnter your Meta Pixel ID"),
    ]),
    ("EVERMORE LIFE SALES MATERIALS", [
        ("38", "Agent binder — 9-section Word document", STATUS_FILE,
         "File: Evermore_Life_Agent_Binder.docx\nComplete: scripts, objections, carrier info, IUL guide, daily scorecard"),
        ("39", "Lead tracker spreadsheet", STATUS_FILE,
         "File: Evermore_Life_Lead_Tracker.xlsx\nComplete: 4 sheets, color-coded status, chart, 52 formulas"),
        ("40", "Final expense ad funnel (existing files)", STATUS_YOU,
         "YOU DO IT (home Mac):\nOpen Cowork → select Evermore Life project folder\nReview existing landing page design\nTell Claude what needs to change"),
        ("41", "Master Funnel Playbook (strategy + all ad sets)", STATUS_FILE,
         "File: Evermore_Meta_Funnel_Playbook.xlsx\n6 tabs: Funnel Map, Strategy, Ad Sets, Scripts, Image Prompts, KPIs"),
        ("42", "Project Playbook (roadmap doc)", STATUS_FILE,
         "File: Evermore_Life_Playbook.docx\nComplete overview of all projects, status, and next steps"),
    ]),
    ("COMPLIANCE & LEGAL", [
        ("43", "Ad compliance language added to all scripts", STATUS_DONE,
         "Included in all scripts: 'Not affiliated with any government program. Licensed insurance agent. Coverage and rates vary by age and health.'\nSee Evermore_Video_Ad_Scripts.docx — Compliance page."),
        ("44", "Sarah conversation — no specific coverage promises", STATUS_DONE,
         "Sarah's script never promises specific dollar amounts. Language: 'Most of our clients in your situation qualify for...' (ranges only)"),
        ("45", "State licensing verification", STATUS_YOU,
         "YOU DO IT:\nVerify you are licensed and in good standing in every state you intend to advertise\nIn GHL: add your NPN (National Producer Number) to your agent profile\nCheck NIPR.com for license status"),
        ("46", "Privacy policy for Sarah's page", STATUS_YOU,
         "YOU DO IT:\nAdd a privacy policy link to Sarah's page footer\nUse Termly.io (free) to generate one\nMeta requires privacy policy for lead ads"),
    ]),
]

current_row = 6
for section, items in tasks:
    # Section header
    ws.merge_cells(f"B{current_row}:E{current_row}")
    ws[f"B{current_row}"].value = section
    ws[f"B{current_row}"].fill = fill(NAVY)
    ws[f"B{current_row}"].font = Font(bold=True, size=11, color=GOLD, name="Arial")
    ws[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    for num, task, (status_txt, status_bg, status_fg), instructions in items:
        # Number
        ws[f"B{current_row}"].value = num
        ws[f"B{current_row}"].fill = fill(LNAVY)
        ws[f"B{current_row}"].font = Font(bold=True, size=9, color=NAVY, name="Arial")
        ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{current_row}"].border = bdr()

        # Task name
        ws[f"C{current_row}"].value = task
        ws[f"C{current_row}"].fill = fill(LGRAY)
        ws[f"C{current_row}"].font = Font(size=10, color="111111", name="Arial")
        ws[f"C{current_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"C{current_row}"].border = bdr()

        # Status
        ws[f"D{current_row}"].value = status_txt
        ws[f"D{current_row}"].fill = fill(status_bg)
        ws[f"D{current_row}"].font = Font(bold=True, size=10, color=status_fg, name="Arial")
        ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"D{current_row}"].border = bdr()

        # Instructions
        ws[f"E{current_row}"].value = instructions
        ws[f"E{current_row}"].fill = fill(WHITE)
        ws[f"E{current_row}"].font = Font(size=9, color="333333", name="Arial")
        ws[f"E{current_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{current_row}"].border = bdr()

        lines = instructions.count("\n") + 1
        ws.row_dimensions[current_row].height = max(28, min(lines * 14, 90))
        current_row += 1

    current_row += 1

ws.freeze_panes = "B6"

PROJECT_ROOT = Path(__file__).resolve().parents[2]
out = PROJECT_ROOT / "03_sales_marketing" / "playbooks" / "Evermore_Implementation_Tracker.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out))
print(f"Saved: {out}")
